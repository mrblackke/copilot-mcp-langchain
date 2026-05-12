import mcp.types as types
import os
import json
import base64
import requests
from collections import defaultdict

# ---------------------------------------------------------------------------
# Role detection constants (shared with worklog_report)
# ---------------------------------------------------------------------------
ROLES = ["frontend", "backend", "qa", "ba", "other"]

ROLE_KEYWORDS = {
    "frontend": ["fe:", "fe ", "[fe]", "(fe)", "frontend", "front-end", "ui:", "ui ", "[ui]", "(ui)", "react", "vue", "angular", "css", "html"],
    "backend":  ["be:", "be ", "[be]", "(be)", "backend", "back-end", "api:", "api ", "[api]", "(api)", "server:", "service:", "db:", "database"],
    "qa":       ["qa:", "qa ", "[qa]", "(qa)", "qc:", "qc ", "[qc]", "(qc)", "test:", "testing", "autotests", "e2e:", "manual:"],
    "ba":       ["ba:", "ba ", "[ba]", "(ba)", "analyst", "analysis", "requirement", "requirements", "business analysis", "grooming", "backlog", "acceptance criteria"],
}

# Exact single-word subtask names used by EPMBWF team (e.g. subtask summary = "BE", "FE", "QA")
ROLE_EXACT = {
    "fe": "frontend",
    "be": "backend",
    "qa": "qa",
    "ba": "ba",
    "frontend": "frontend",
    "backend": "backend",
}

# ---------------------------------------------------------------------------
# Team roster: name_substring (lowercase) → role
# Defines who is on the team and what role each person plays.
# Used for auto-assignment: tasks are assigned to least-loaded person by role.
# Add new team members here or pass them via the `team_roles` parameter.
# ---------------------------------------------------------------------------
TEAM_ROSTER: dict[str, str] = {
    # BA
    "darya strogonova": "ba",
    "elvina": "ba",
    # Frontend
    "anton pankou": "frontend",
    "aleksandr stromov": "frontend",
    "andrei hrakovich": "frontend",
    # Backend
    "ruslan zhyvotovskyi": "backend",
    "daulet kshibekov": "backend",
    "yegor osipov": "backend",
    # QA
    "yelena zhuravleva": "qa",
    "vitali talipski": "qa",
    "iryna shaldaieva": "qa",
    "yauheni basalyha": "qa",
}

# Alias used by _detect_role (keeps worklog_report parity)
PERSON_ROLE_OVERRIDES = TEAM_ROSTER


def _secs_to_h(s: int | float) -> float:
    return round(s / 3600, 2)


def _detect_role(summary: str, author_name: str = "") -> str:
    """Detect task role from summary keywords. Person override takes priority."""
    if author_name:
        name_lower = author_name.lower()
        for person_substr, role in PERSON_ROLE_OVERRIDES.items():
            if person_substr in name_lower:
                return role
    s = summary.strip().lower()
    # Exact match first (handles single-word subtask names like "BE", "FE", "QA")
    if s in ROLE_EXACT:
        return ROLE_EXACT[s]
    for role, keywords in ROLE_KEYWORDS.items():
        if any(k in s for k in keywords):
            return role
    return "other"


def _make_headers(jira_session: str = "", jira_username: str = "",
                  jira_password: str = "", jira_auth: str = "") -> dict:
    headers = {"Accept": "application/json"}
    if jira_session:
        headers["Cookie"] = jira_session
    elif jira_username and jira_password:
        token = base64.b64encode(f"{jira_username}:{jira_password}".encode()).decode()
        headers["Authorization"] = f"Basic {token}"
    elif jira_auth:
        headers["Authorization"] = f"Bearer {jira_auth}"
    return headers


def _get_board_id(session: requests.Session, jira_url: str, project: str) -> int | None:
    resp = session.get(
        f"{jira_url}/rest/agile/1.0/board",
        params={"projectKeyOrId": project, "maxResults": 50},
        timeout=30,
    )
    resp.raise_for_status()
    values = resp.json().get("values", [])
    if not values:
        return None
    scrum = [b for b in values if b.get("type", "").lower() == "scrum"]
    return scrum[0]["id"] if scrum else values[0]["id"]


def _get_sprint_by_state(session: requests.Session, jira_url: str,
                          board_id: int, state: str) -> dict | None:
    """Return the first sprint with given state (active → current, future → next)."""
    resp = session.get(
        f"{jira_url}/rest/agile/1.0/board/{board_id}/sprint",
        params={"state": state, "startAt": 0, "maxResults": 50},
        timeout=30,
    )
    resp.raise_for_status()
    sprints = resp.json().get("values", [])
    return sprints[0] if sprints else None


def _get_all_sprint_issues(session: requests.Session, jira_url: str, sprint_id: int) -> list:
    """Paginated fetch of all sprint issues with fields needed for planning."""
    issues = []
    start = 0
    fields = "summary,timeoriginalestimate,assignee,issuetype,parent,status,subtasks"
    while True:
        resp = session.get(
            f"{jira_url}/rest/agile/1.0/sprint/{sprint_id}/issue",
            params={"startAt": start, "maxResults": 100, "fields": fields},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        batch = data.get("issues", [])
        issues.extend(batch)
        if start + 100 >= data.get("total", 0):
            break
        start += 100
    return issues


def _resolve_planning_issues(all_issues: list) -> tuple[list, list]:
    """
    Handle story/subtask hierarchy for estimation:
    - If a story has subtasks in this sprint → use subtask estimates, skip story
    - If a story has no subtasks → use story estimate
    - Direct tasks/bugs → always include

    Returns: (estimated_issues, unestimated_issues)
    """
    sprint_keys = {i["key"] for i in all_issues}
    skip_keys: set[str] = set()

    # Any issue whose parent is also in this sprint → parent is a story with subtasks
    for issue in all_issues:
        parent_key = (issue.get("fields", {}).get("parent") or {}).get("key", "")
        if parent_key in sprint_keys:
            skip_keys.add(parent_key)

    estimated = []
    unestimated = []
    for issue in all_issues:
        if issue["key"] in skip_keys:
            continue
        est = (issue.get("fields") or {}).get("timeoriginalestimate") or 0
        if est > 0:
            estimated.append(issue)
        else:
            unestimated.append(issue)

    return estimated, unestimated


def _get_carryover(session: requests.Session, jira_url: str, board_id: int) -> tuple[dict | None, list]:
    """Returns (active_sprint, unfinished_issues). Empty list if no active sprint."""
    active = _get_sprint_by_state(session, jira_url, board_id, "active")
    if not active:
        return None, []

    issues = _get_all_sprint_issues(session, jira_url, active["id"])
    done_statuses = {"done", "closed", "resolved", "won't fix", "wont fix", "cancelled"}
    carryover = []
    for issue in issues:
        fields = issue.get("fields", {})
        status_name = (fields.get("status", {}).get("name") or "").lower()
        if status_name not in done_statuses:
            assignee = (fields.get("assignee") or {}).get("displayName", "Unassigned")
            est = fields.get("timeoriginalestimate") or 0
            carryover.append({
                "issue_key": issue["key"],
                "summary": fields.get("summary", ""),
                "type": (fields.get("issuetype", {}).get("name") or ""),
                "status": (fields.get("status", {}).get("name") or "Unknown"),
                "assignee": assignee,
                "estimated_hours": _secs_to_h(est),
            })
    return active, carryover


def _merge_team_roles(extra_roles: dict) -> dict[str, str]:
    """Merge TEAM_ROSTER with extra_roles param (param takes priority)."""
    merged = dict(TEAM_ROSTER)
    for name, role in extra_roles.items():
        merged[name.lower().strip()] = role.lower().strip()
    return merged


def _resolve_vacation_hours(name_key: str, vacations: dict, default_capacity: float) -> float:
    """Find vacation hours for a person by case-insensitive substring match."""
    for vac_name, vac_h in vacations.items():
        vac_lower = vac_name.lower().strip()
        if vac_lower in name_key or name_key in vac_lower:
            return float(vac_h)
    return default_capacity


def _assign_plan(
    estimated_issues: list,
    team_roles: dict[str, str],
    capacity_per_person: float,
    vacations: dict,
    overload_threshold: float = 110.0,
) -> tuple[dict, list]:
    """
    Auto-assign estimated tasks to team members by role (least-loaded first).

    Returns:
      plan: {name_key → {display_name, role, capacity_h, planned_h, remaining_h, load_pct, status, tasks}}
      no_role_warnings: tasks that couldn't be assigned (role='other' or no matching team member)
    """
    # Build team registry
    team: dict[str, dict] = {}
    for name_key, role in team_roles.items():
        cap_h = _resolve_vacation_hours(name_key, vacations, capacity_per_person)
        team[name_key] = {
            "display_name": name_key.title(),
            "role": role,
            "capacity_h": cap_h,
            "planned_h": 0.0,
            "tasks": [],
        }

    # Group team members by role for fast lookup
    members_by_role: dict[str, list[str]] = defaultdict(list)
    for name_key, info in team.items():
        members_by_role[info["role"]].append(name_key)

    no_role_warnings: list[dict] = []

    for issue in estimated_issues:
        fields = issue.get("fields", {})
        summary = fields.get("summary", "")
        est_h = _secs_to_h(fields.get("timeoriginalestimate", 0))
        jira_assignee = (fields.get("assignee") or {}).get("displayName", "")

        # Detect task role from summary only (no author context in planning)
        task_role = _detect_role(summary)

        candidates = members_by_role.get(task_role, [])
        if not candidates:
            no_role_warnings.append({
                "issue_key": issue["key"],
                "summary": summary,
                "detected_role": task_role,
                "estimated_hours": est_h,
                "jira_assignee": jira_assignee,
                "reason": f"No team members with role '{task_role}' in team roster",
            })
            continue

        # Assign to least-loaded team member with matching role
        best = min(candidates, key=lambda k: team[k]["planned_h"])
        team[best]["planned_h"] = round(team[best]["planned_h"] + est_h, 2)
        team[best]["tasks"].append({
            "issue_key": issue["key"],
            "summary": summary,
            "role": task_role,
            "estimated_hours": est_h,
            "jira_assignee": jira_assignee,
        })

    # Build final plan with status emoji
    plan: dict[str, dict] = {}
    for name_key, info in team.items():
        cap = info["capacity_h"]
        planned = info["planned_h"]
        pct = round(planned / cap * 100, 1) if cap > 0 else 0.0
        if pct > overload_threshold:
            status = "🔴 overloaded"
        elif pct >= 90:
            status = "🟢 balanced"
        else:
            status = "🟡 underloaded"
        plan[name_key] = {
            "display_name": info["display_name"],
            "role": info["role"],
            "capacity_h": cap,
            "planned_h": planned,
            "remaining_h": round(cap - planned, 2),
            "load_pct": pct,
            "status": status,
            "tasks": info["tasks"],
        }

    return plan, no_role_warnings


def _build_role_summary(plan: dict, overload_threshold: float = 110.0) -> list:
    """Aggregate capacity and load per role for the role summary sheet."""
    by_role: dict[str, dict] = defaultdict(lambda: {"capacity_h": 0.0, "planned_h": 0.0, "members": 0})
    for info in plan.values():
        role = info["role"]
        by_role[role]["capacity_h"] += info["capacity_h"]
        by_role[role]["planned_h"] += info["planned_h"]
        by_role[role]["members"] += 1

    result = []
    for role in ROLES:
        if role not in by_role:
            continue
        d = by_role[role]
        pct = round(d["planned_h"] / d["capacity_h"] * 100, 1) if d["capacity_h"] > 0 else 0.0
        result.append({
            "role": role,
            "members": d["members"],
            "capacity_h": round(d["capacity_h"], 1),
            "planned_h": round(d["planned_h"], 1),
            "remaining_h": round(d["capacity_h"] - d["planned_h"], 1),
            "load_pct": pct,
            "status": "🔴 overloaded" if pct > overload_threshold else ("🟢 balanced" if pct >= 90 else "🟡 underloaded"),
        })
    return result


def _build_excel(
    plan: dict,
    unestimated: list,
    no_role_warnings: list,
    carryover: list,
    role_summary: list,
    future_sprint: dict,
    output_path: str,
) -> str:
    """Generate single-sheet Excel matching the BWF sprint planning template layout.

    Sections (top → bottom):
      1. Sprint Capacity (person table)
      2. Role Summary (FE/BE/QA/BA totals)
      3. Committed Tasks
      4. ⚠️ Unestimated tasks  (if any)
      5. ⚠️ No role match      (if any)
      6. Carryover from active sprint
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install it: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Plan"
    sprint_name = future_sprint.get("name", "Next Sprint")

    # ── Column widths (fixed once, all sections share the same 6 columns) ──
    COL_WIDTHS = [12, 52, 14, 10, 26, 17]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Shared styles ────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(color="FFFFFF", bold=True)
    SEC_FONT  = Font(bold=True, size=11, color="1F3864")
    RED_FILL  = PatternFill("solid", fgColor="FFD7D7")
    GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")
    YEL_FILL  = PatternFill("solid", fgColor="FFF2CC")
    ORG_FILL  = PatternFill("solid", fgColor="FCE4D6")
    WARN_FONT = Font(bold=True, color="CC0000")
    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="1F3864")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    NCOLS = 6  # number of columns used throughout

    def _add_gap():
        ws.append([""] * NCOLS)
        ws.row_dimensions[ws.max_row].height = 8

    def _section_title(text: str):
        _add_gap()
        ws.append([text])
        row = ws.max_row
        cell = ws.cell(row=row, column=1)
        cell.font = SEC_FONT
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
        cell.border = Border(bottom=medium)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        ws.row_dimensions[row].height = 20

    def _header_row(labels: list[str]):
        ws.append(labels)
        row = ws.max_row
        for i, _ in enumerate(labels, 1):
            cell = ws.cell(row=row, column=i)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 20

    def _data_row(values: list, fill=None):
        # Pad / trim to NCOLS so borders are consistent
        padded = (values + [""] * NCOLS)[:NCOLS]
        ws.append(padded)
        row = ws.max_row
        for i in range(1, NCOLS + 1):
            cell = ws.cell(row=row, column=i)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
        ws.row_dimensions[row].height = 18

    def _warn_label(text: str):
        ws.append([text])
        row = ws.max_row
        ws.cell(row=row, column=1).font = WARN_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        ws.row_dimensions[row].height = 18

    # ── Sprint title ─────────────────────────────────────────────────────────
    ws.append([f"Sprint Plan — {sprint_name}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    ws.row_dimensions[1].height = 26

    role_order = {r: i for i, r in enumerate(ROLES)}

    # ═════════════════════════════════════════════════════════════════════════
    # Section 1: SPRINT CAPACITY
    # ═════════════════════════════════════════════════════════════════════════
    _section_title("SPRINT CAPACITY")
    _header_row(["Person", "Role", "Capacity (h)", "Planned (h)", "Remaining (h)", "Status"])
    for name_key, info in sorted(plan.items(), key=lambda x: role_order.get(x[1]["role"], 99)):
        fill = RED_FILL if "overloaded" in info["status"] else (GRN_FILL if "balanced" in info["status"] else YEL_FILL)
        _data_row([
            info["display_name"],
            info["role"].upper(),
            info["capacity_h"],
            info["planned_h"],
            info["remaining_h"],
            info["status"],
        ], fill=fill)

    # ── Totals row ────────────────────────────────────────────────────────────
    total_cap     = round(sum(i["capacity_h"] for i in plan.values()), 1)
    total_planned = round(sum(i["planned_h"]  for i in plan.values()), 1)
    total_rem     = round(total_cap - total_planned, 1)
    ws.append(["TOTAL", "", total_cap, total_planned, total_rem, ""])
    row = ws.max_row
    for i in range(1, NCOLS + 1):
        cell = ws.cell(row=row, column=i)
        cell.font = Font(bold=True)
        cell.border = Border(top=medium, bottom=medium, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18

    # ═════════════════════════════════════════════════════════════════════════
    # Section 2: ROLE SUMMARY
    # ═════════════════════════════════════════════════════════════════════════
    _section_title("ROLE SUMMARY")
    _header_row(["Role", "Members", "Capacity (h)", "Planned (h)", "Remaining (h)", "Load %"])
    for rs in role_summary:
        fill = RED_FILL if "overloaded" in rs["status"] else (GRN_FILL if "balanced" in rs["status"] else YEL_FILL)
        _data_row([
            rs["role"].upper(),
            rs["members"],
            rs["capacity_h"],
            rs["planned_h"],
            rs["remaining_h"],
            f"{rs['load_pct']}%  {rs['status']}",
        ], fill=fill)

    # ═════════════════════════════════════════════════════════════════════════
    # Section 3: COMMITTED TASKS
    # ═════════════════════════════════════════════════════════════════════════
    _section_title("COMMITTED TASKS")
    _header_row(["Issue", "Summary", "Role", "Est. (h)", "Assigned To", "Jira Assignee"])
    for name_key, info in sorted(plan.items(), key=lambda x: role_order.get(x[1]["role"], 99)):
        for task in info["tasks"]:
            _data_row([
                task["issue_key"],
                task["summary"],
                task["role"].upper(),
                task["estimated_hours"],
                info["display_name"],
                task.get("jira_assignee") or "—",
            ])

    # ═════════════════════════════════════════════════════════════════════════
    # Section 4: UNESTIMATED TASKS (if any)
    # ═════════════════════════════════════════════════════════════════════════
    if unestimated:
        _add_gap()
        _warn_label(f"⚠️  UNESTIMATED TASKS — {len(unestimated)} task(s) excluded from planning (add time estimates in Jira)")
        _header_row(["Issue", "Summary", "Type", "Status", "Jira Assignee", ""])
        for issue in unestimated:
            _data_row([
                issue["issue_key"],
                issue["summary"],
                issue.get("type", ""),
                issue.get("status", ""),
                issue.get("jira_assignee", "—"),
                "",
            ], fill=YEL_FILL)

    # ═════════════════════════════════════════════════════════════════════════
    # Section 5: NO ROLE MATCH (if any)
    # ═════════════════════════════════════════════════════════════════════════
    if no_role_warnings:
        _add_gap()
        _warn_label(f"⚠️  NO ROLE MATCH — {len(no_role_warnings)} task(s) not auto-assigned (add [FE]/[BE]/[QA] tag to summary)")
        _header_row(["Issue", "Summary", "Detected Role", "Est. (h)", "Reason", ""])
        for w in no_role_warnings:
            _data_row([
                w["issue_key"],
                w["summary"],
                w["detected_role"],
                w["estimated_hours"],
                w["reason"],
                "",
            ], fill=ORG_FILL)

    # ═════════════════════════════════════════════════════════════════════════
    # Section 6: CARRYOVER
    # ═════════════════════════════════════════════════════════════════════════
    _section_title("CARRYOVER — Unfinished Tasks from Active Sprint")
    if carryover:
        _header_row(["Issue", "Summary", "Type", "Status", "Assignee", "Est. (h)"])
        for item in carryover:
            _data_row([
                item["issue_key"],
                item["summary"],
                item.get("type", ""),
                item["status"],
                item["assignee"],
                item["estimated_hours"],
            ], fill=YEL_FILL)
    else:
        ws.append(["✅  No carryover — all active sprint tasks are Done/Closed."])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="006100")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=NCOLS)
        ws.row_dimensions[ws.max_row].height = 18

    wb.save(output_path)
    return output_path


def _text_summary(
    plan: dict,
    role_summary: list,
    carryover: list,
    unestimated: list,
    no_role_warnings: list,
    future_sprint: dict,
) -> str:
    sprint_name = future_sprint.get("name", "Next Sprint")
    total_planned = sum(info["planned_h"] for info in plan.values())
    total_capacity = sum(info["capacity_h"] for info in plan.values())

    lines = [
        f"## Sprint Plan: {sprint_name}",
        "",
        f"**Carryover:** {len(carryover)} unfinished tasks from active sprint",
        f"**Team capacity:** {total_capacity:.0f}h total | {total_planned:.1f}h planned "
        f"({round(total_planned / total_capacity * 100, 1) if total_capacity else 0}% loaded)",
        "",
        "### Capacity by person",
        f"{'Person':<28} {'Role':<10} {'Cap':>5} {'Plan':>6} {'Rem':>6}  Status",
        "─" * 68,
    ]
    role_order = {r: i for i, r in enumerate(ROLES)}
    for name_key, info in sorted(plan.items(), key=lambda x: role_order.get(x[1]["role"], 99)):
        lines.append(
            f"{info['display_name']:<28} {info['role'].upper():<10} "
            f"{info['capacity_h']:>5.0f} {info['planned_h']:>6.1f} "
            f"{info['remaining_h']:>6.1f}  {info['status']}"
        )

    if role_summary:
        lines += ["", "### Role summary"]
        for rs in role_summary:
            lines.append(
                f"  {rs['role'].upper():<12} {rs['members']} people | "
                f"{rs['planned_h']:.0f}h / {rs['capacity_h']:.0f}h ({rs['load_pct']}%) {rs['status']}"
            )

    if unestimated:
        lines += ["", f"⚠️  **{len(unestimated)} unestimated tasks** — add time estimates to include them in planning:"]
        for issue in unestimated[:5]:
            f = issue.get("fields", {})
            lines.append(f"  - {issue['key']}: {f.get('summary', '')[:70]}")
        if len(unestimated) > 5:
            lines.append(f"  … and {len(unestimated) - 5} more (see Excel COMMITTED TASKS sheet)")

    if no_role_warnings:
        lines += ["", f"⚠️  **{len(no_role_warnings)} tasks not auto-assigned** (role not matched):"]
        for w in no_role_warnings[:5]:
            lines.append(f"  - {w['issue_key']}: {w['summary'][:60]} [detected: {w['detected_role']}]")
        if len(no_role_warnings) > 5:
            lines.append(f"  … and {len(no_role_warnings) - 5} more")

    if carryover:
        lines += ["", f"⚠️  **Carryover ({len(carryover)} tasks):**"]
        for item in carryover[:5]:
            lines.append(f"  - {item['issue_key']}: {item['summary'][:60]} [{item['status']}] → {item['assignee']}")
        if len(carryover) > 5:
            lines.append(f"  … and {len(carryover) - 5} more (see Excel CARRYOVER sheet)")

    return "\n".join(lines)


async def tool_info() -> dict:
    return {
        "description": """Plans the next Jira sprint: reads future sprint tasks, auto-assigns them by role to least-loaded team member, checks carryover from active sprint, and outputs a capacity plan as JSON + Excel.

**How it works:**
1. **Phase 0 (Carryover):** Checks active sprint for unfinished tasks → warns about carryover risk.
2. **Phase 1 (Future sprint):** Fetches all issues from the next (future) sprint already created in Jira.
3. **Phase 2 (Assignment):** Detects each task's role from summary ([FE]/[BE]/[QA] keywords), then assigns it to the least-loaded team member with that role.
4. **Phase 3 (Overload check):** Flags overloaded (>110%) / balanced (90–110%) / underloaded (<90%) people.
5. **Output:** Text summary in chat + JSON file + Excel file (single sheet, sections top→bottom).

**Excel sections (one sheet):**
- SPRINT CAPACITY — person table with capacity / planned / remaining + TOTAL row
- ROLE SUMMARY — FE/BE/QA/BA aggregate load %
- COMMITTED TASKS — full task list with auto-assignment
- ⚠️ UNESTIMATED TASKS — tasks excluded from planning (if any)
- ⚠️ NO ROLE MATCH — tasks not auto-assigned (if any)
- CARRYOVER — unfinished tasks from active sprint

**Team roster** is defined in TEAM_ROSTER inside tool.py. Extend it or use `team_roles` param to add/override members dynamically.

**Parameters:**
- `project` (string, required): Jira project key e.g. "EPMBWF"
- `board_id` (int, optional): Scrum board ID — skips auto-discovery. Use when project has multiple boards.
- `capacity_hours_per_person` (float, optional): Default capacity per person in hours (default: 90 = 6h/day × 15 days).
- `vacations` (string, optional): JSON string to override capacity for specific people. E.g. `{"Anton Pankou": 60, "Darya Strogonova": 45}`.
- `team_roles` (string, optional): JSON string to add/override person→role mapping. E.g. `{"new person": "frontend", "another": "qa"}`.
- `overload_threshold` (float, optional): % of capacity for overload warning (default: 110).
- `output_file` (string, optional): Base path for output files. If provided, saves `<output_file>.json` and `<output_file>.xlsx`.
- `env_file` (string, optional): Path to .env file (default: ".env").

**Environment Variables Required:**
- `JIRA_URL`: Base URL of your Jira instance (e.g., "https://jiraeu.epam.com")

Auth options (first match wins):
1. `JIRA_SESSION` — browser session cookie (recommended for SSO/Microsoft login)
2. `JIRA_USERNAME` + `JIRA_PASSWORD` — Basic auth
3. `JIRA_AUTH` — Bearer/Personal Access Token

**Example Usage:**
- Basic plan:   `{"project": "EPMBWF"}`
- With vacations: `{"project": "EPMBWF", "vacations": "{\\"Anton Pankou\\": 60}"}`
- Save Excel: `{"project": "EPMBWF", "output_file": "sprint_plan"}`
- Explicit board: `{"project": "EPMBWF", "board_id": 240635, "output_file": "sprint_plan"}`""",
        "schema": {
            "type": "object",
            "properties": {
                "project": {
                    "type": "string",
                    "description": "Jira project key (e.g., EPMBWF)"
                },
                "board_id": {
                    "type": "integer",
                    "description": "Optional: Jira board ID — skips auto-discovery"
                },
                "capacity_hours_per_person": {
                    "type": "number",
                    "description": "Default capacity per person in hours (default: 90)"
                },
                "vacations": {
                    "type": "string",
                    "description": "JSON string: {\"Person Name\": hours} — overrides default capacity"
                },
                "team_roles": {
                    "type": "string",
                    "description": "JSON string: {\"person name substring\": \"role\"} — adds/overrides team roster"
                },
                "overload_threshold": {
                    "type": "number",
                    "description": "% of capacity threshold for overload warning (default: 110)"
                },
                "output_file": {
                    "type": "string",
                    "description": "Base path for output — saves <output_file>.json and <output_file>.xlsx"
                },
                "env_file": {
                    "type": "string",
                    "description": "Path to .env file (default: .env)"
                }
            },
            "required": ["project"]
        }
    }


async def run_tool(name: str, parameters: dict) -> list[types.Content]:
    """Plans the next Jira sprint with auto-assignment and capacity analysis."""
    try:
        try:
            from dotenv import load_dotenv
        except ImportError:
            return [types.TextContent(type="text", text=json.dumps({
                "success": False,
                "error": "Required dependency not available: python-dotenv"
            }, indent=2))]

        project = parameters.get("project", "").strip()
        if not project:
            return [types.TextContent(type="text", text=json.dumps({
                "success": False, "error": "project parameter is required"
            }, indent=2))]

        capacity_per_person = float(parameters.get("capacity_hours_per_person", 90))
        overload_threshold = float(parameters.get("overload_threshold", 110))
        output_file = parameters.get("output_file")
        env_file = parameters.get("env_file", ".env")

        # Parse JSON params
        try:
            vacations_raw = parameters.get("vacations", "{}")
            vacations: dict = json.loads(vacations_raw) if isinstance(vacations_raw, str) else (vacations_raw or {})
        except (json.JSONDecodeError, TypeError):
            return [types.TextContent(type="text", text=json.dumps({
                "success": False, "error": "vacations must be a valid JSON string, e.g. '{\"Person\": 60}'"
            }, indent=2))]

        try:
            team_roles_raw = parameters.get("team_roles", "{}")
            team_roles_extra: dict = json.loads(team_roles_raw) if isinstance(team_roles_raw, str) else (team_roles_raw or {})
        except (json.JSONDecodeError, TypeError):
            return [types.TextContent(type="text", text=json.dumps({
                "success": False, "error": "team_roles must be a valid JSON string"
            }, indent=2))]

        # Load .env
        env_path = os.path.expanduser(env_file)
        if not os.path.isabs(env_path):
            env_path = os.path.abspath(env_path)
        if not os.path.exists(env_path):
            return [types.TextContent(type="text", text=json.dumps({
                "success": False, "error": f".env file not found at {env_path}"
            }, indent=2))]
        load_dotenv(env_path)

        jira_url = os.getenv("JIRA_URL", "").rstrip("/")
        jira_session = os.getenv("JIRA_SESSION", "")
        jira_username = os.getenv("JIRA_USERNAME", "")
        jira_password = os.getenv("JIRA_PASSWORD", "")
        jira_auth = os.getenv("JIRA_AUTH", "")

        if not jira_url:
            return [types.TextContent(type="text", text=json.dumps({
                "success": False, "error": "JIRA_URL not found in environment variables"
            }, indent=2))]
        if not jira_session and not (jira_username and jira_password) and not jira_auth:
            return [types.TextContent(type="text", text=json.dumps({
                "success": False,
                "error": (
                    "No Jira credentials found. Set one of:\n"
                    "1. JIRA_SESSION — browser session cookie (use lng_jira_sso_auth to get it)\n"
                    "2. JIRA_USERNAME + JIRA_PASSWORD — Basic auth\n"
                    "3. JIRA_AUTH — Bearer/PAT token"
                )
            }, indent=2))]

        with requests.Session() as session:
            session.headers.update(
                _make_headers(jira_session, jira_username, jira_password, jira_auth)
            )

            # --- Step 1: Resolve board ---
            board_id = parameters.get("board_id")
            if board_id:
                board_id = int(board_id)
            else:
                board_id = _get_board_id(session, jira_url, project)
            if board_id is None:
                return [types.TextContent(type="text", text=json.dumps({
                    "success": False,
                    "error": f"No board found for project '{project}'. Check the project key."
                }, indent=2))]

            # --- Step 2: Carryover from active sprint ---
            active_sprint, carryover = _get_carryover(session, jira_url, board_id)

            # --- Step 3: Future sprint ---
            future_sprint = _get_sprint_by_state(session, jira_url, board_id, "future")
            if future_sprint is None:
                return [types.TextContent(type="text", text=json.dumps({
                    "success": False,
                    "error": (
                        "No future sprint found on this board. "
                        "Please create the next sprint in Jira before running the planner."
                    )
                }, indent=2))]

            # --- Step 4: Fetch and resolve future sprint issues ---
            all_future_issues = _get_all_sprint_issues(session, jira_url, future_sprint["id"])

        if not all_future_issues:
            return [types.TextContent(type="text", text=json.dumps({
                "success": False,
                "error": (
                    f"Future sprint '{future_sprint['name']}' has no issues. "
                    "Add tasks to the sprint before planning."
                )
            }, indent=2))]

        estimated, unestimated = _resolve_planning_issues(all_future_issues)

        # --- Step 5: Merge team roles ---
        team_roles = _merge_team_roles(team_roles_extra)

        # --- Step 6: Auto-assign ---
        plan, no_role_warnings = _assign_plan(
            estimated, team_roles, capacity_per_person, vacations, overload_threshold
        )

        # --- Step 7: Role summary ---
        role_summary = _build_role_summary(plan, overload_threshold)

        # --- Step 8: Build result JSON ---
        result = {
            "success": True,
            "operation": "lng_jira_sprint_planner",
            "project": project,
            "board_id": board_id,
            "future_sprint": {
                "id": future_sprint["id"],
                "name": future_sprint["name"],
                "start_date": (future_sprint.get("startDate") or "")[:10],
                "end_date": (future_sprint.get("endDate") or "")[:10],
            },
            "active_sprint": {
                "id": active_sprint["id"] if active_sprint else None,
                "name": active_sprint["name"] if active_sprint else None,
            },
            "capacity_per_person_h": capacity_per_person,
            "overload_threshold_pct": overload_threshold,
            "carryover_count": len(carryover),
            "carryover": carryover,
            "plan": plan,
            "role_summary": role_summary,
            "total_estimated_tasks": len(estimated),
            "total_planned_hours": round(sum(i["planned_h"] for i in plan.values()), 1),
            "total_capacity_hours": round(sum(i["capacity_h"] for i in plan.values()), 1),
            "unestimated_warnings": [
                {
                    "issue_key": i["key"],
                    "summary": (i.get("fields") or {}).get("summary", ""),
                    "type": ((i.get("fields") or {}).get("issuetype") or {}).get("name", ""),
                    "status": ((i.get("fields") or {}).get("status") or {}).get("name", ""),
                    "jira_assignee": ((i.get("fields") or {}).get("assignee") or {}).get("displayName", "Unassigned"),
                }
                for i in unestimated
            ],
            "no_role_warnings": no_role_warnings,
        }

        # --- Step 9: Save output files ---
        excel_path = None
        if output_file:
            base_path = os.path.expanduser(output_file)
            if not os.path.isabs(base_path):
                base_path = os.path.abspath(base_path)
            # Strip extension if user passed one
            if base_path.endswith(".json") or base_path.endswith(".xlsx"):
                base_path = base_path[:-5]
            # Append sprint name: e.g. sprint_plan_EPM-BWF-_Sprint_77
            sprint_suffix = future_sprint.get("name", "").replace(" ", "_").replace("/", "-")
            base_path = f"{base_path}_{sprint_suffix}"
            dir_name = os.path.dirname(base_path)
            if dir_name:
                os.makedirs(dir_name, exist_ok=True)

            # JSON
            json_path = base_path + ".json"
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            result["json_output_file"] = json_path

            # Excel — pass serialized unestimated list (clean dicts, not raw Jira objects)
            excel_path = base_path + ".xlsx"
            _build_excel(
                plan=plan,
                unestimated=result["unestimated_warnings"],
                no_role_warnings=no_role_warnings,
                carryover=carryover,
                role_summary=role_summary,
                future_sprint=future_sprint,
                output_path=excel_path,
            )
            result["excel_output_file"] = excel_path

        # --- Step 10: Build text summary for chat ---
        summary_text = _text_summary(
            plan=plan,
            role_summary=role_summary,
            carryover=carryover,
            unestimated=unestimated,
            no_role_warnings=no_role_warnings,
            future_sprint=future_sprint,
        )
        if excel_path:
            summary_text += f"\n\n📊 **Excel saved:** `{excel_path}`"

        return [
            types.TextContent(type="text", text=summary_text),
            types.TextContent(type="text", text=json.dumps(result, indent=2, ensure_ascii=False)),
        ]

    except requests.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else "?"
        body = ""
        if e.response is not None:
            try:
                body = e.response.json().get("errorMessages", e.response.text[:200])
            except Exception:
                body = e.response.text[:200]
        msg = f"Jira API error {status}: {body}"
        if status == 401:
            msg += "\n→ Session cookie may be expired. Run lng_jira_sso_auth to refresh JIRA_SESSION."
        elif status == 403:
            msg += "\n→ No permission to access this board/project."
        return [types.TextContent(type="text", text=json.dumps({
            "success": False, "error": msg
        }, indent=2))]

    except Exception as e:
        return [types.TextContent(type="text", text=json.dumps({
            "success": False,
            "error": f"{type(e).__name__}: {e}"
        }, indent=2))]
